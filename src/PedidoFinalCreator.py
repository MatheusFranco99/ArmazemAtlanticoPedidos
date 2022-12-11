from datetime import datetime,date
from ShelfLifeReader import *
from UnidadesPorCaixaReader import *
from Day import *
from ExcelLetters import *
from DemandEstimator import *
from ProphetDemandEstimator import *
from ProgressBar import *
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, PatternFill


PEDIDO_FINAL_TEMPLATE_FILE = ".data/PedidoFinalTemplate.xlsx"


class PedidoFinalCreator:


    def __init__(self, filename = 'PedidoFinal.xlsx'):
        
        # process filename
        if filename == PEDIDO_FINAL_TEMPLATE_FILE:
            filename = ".data/PedidoTemplate1.xlsx"

        if filename[-5:] != ".xlsx":
            filename = filename + ".xlsx"
        self.filename = filename

        # copy file from template
        shutil.copyfile(PEDIDO_FINAL_TEMPLATE_FILE,self.filename)

    def addItems(self,items,pedidosEmCaixa):
        
        # load file
        workbook = load_workbook(filename=self.filename)
        sheet = workbook['Pedido']
        


        row = 10
        rowStr = str(row)
        for i in range(len(items)):
            item = items[i]
            qtd = 0
            try:
                qtd = pedidosEmCaixa[item.ID]
            except:
                pass
        
            if item.config == None:
                item.config = { \
                'embalagem':None, \
                'codigo_nf_fornecedor':None, \
                'fornecedor':None, \
                'categoria':None, \
                'tradebras_sistem_code':None, \
                'shelf_life':None, \
                'products':None, \
                'hs_code':None, \
                'ncm':None, \
                'c':None, \
                'l':None, \
                'a':None, \
                'm3_unit':None, \
                'embarque':None, \
                'peso_liq_unit':None, \
                'peso_bruto_unit':None \
                }

            # insert row
            sheet['A'+rowStr] = qtd
            sheet['B'+rowStr] = item.Description
            sheet['C'+rowStr] = item.config['embalagem']
            sheet['D'+rowStr] = item.config['codigo_nf_fornecedor']
            sheet['E'+rowStr] = item.config['fornecedor']
            sheet['F'+rowStr] = item.config['categoria']
            sheet['G'+rowStr] = item.config['tradebras_sistem_code']
            sheet['H'+rowStr] = item.config['shelf_life']
            sheet['I'+rowStr] = item.config['products']
            sheet['J'+rowStr] = item.config['hs_code']
            sheet['K'+rowStr] = item.config['ncm']
            try:
                sheet['L'+rowStr] = float(item.config['c'])
            except:
                sheet['L'+rowStr] = item.config['c']
            try:
                sheet['M'+rowStr] = float(item.config['l'])
            except:
                sheet['M'+rowStr] = item.config['l']
            try:
                sheet['N'+rowStr] = float(item.config['a'])
            except:
                sheet['N'+rowStr] = item.config['a']
            try:
                sheet['O'+rowStr] = float(item.config['m3_unit'])
            except:
                sheet['O'+rowStr] = item.config['m3_unit']

            sheet['P'+rowStr] = "=A"+rowStr+"*O"+rowStr
            sheet['Q'+rowStr] = item.config['embarque']
            sheet['R'+rowStr] = item.config['peso_liq_unit']
            sheet['S'+rowStr] = "=A"+rowStr+"*R"+rowStr
            sheet['T'+rowStr] = item.config['peso_bruto_unit']
            sheet['U'+rowStr] = "=A"+rowStr+"*T"+rowStr
            sheet['V'+rowStr] = "=P"+rowStr+"/30"
            sheet['W'+rowStr] = "=P"+rowStr+"/61"

            row = row + 1
            rowStr = str(row)
        
        # add last row
        lastElmRowStr = str(row-1)
        sheet['A'+rowStr] = "=SUM(A10:A"+lastElmRowStr+")"
        sheet['P'+rowStr] = "=SUM(P10:P"+lastElmRowStr+")"
        sheet['S'+rowStr] = "=SUM(S10:S"+lastElmRowStr+")"
        sheet['U'+rowStr] = "=SUM(U10:U"+lastElmRowStr+")"
        sheet['V'+rowStr] = "=SUM(V10:V"+lastElmRowStr+")"
        sheet['W'+rowStr] = "=SUM(W10:W"+lastElmRowStr+")"

        # add header formular
        sheet['B3'] = "=V"+rowStr
        sheet['B4'] = "=W"+rowStr
        sheet['A6'] = "=S"+rowStr
        sheet['B6'] = "=U"+rowStr
        sheet['A8'] = "=P"+rowStr

        # set font color and fill color of last row
        yellowFill = PatternFill(start_color='FFFF00',
                   end_color='FFFF00',
                   fill_type='solid')
        redFont = Font(color="FF0000")

        for letter in ['A','B','C','D','E','F','G','H','I','J','K','L','M','M','N','O','P','Q','R','S','T','U','V','W']:
            sheet[letter + rowStr].fill = yellowFill
            sheet[letter + rowStr].font = redFont
        
        workbook.save(filename = self.filename)
      