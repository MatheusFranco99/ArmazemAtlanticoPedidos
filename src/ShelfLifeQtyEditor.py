from openpyxl import load_workbook

class ShelfLifeQtyEditor:

    @staticmethod
    def update(lst):
        # lst element: [id,description,shelfLife,Qty]
        # shelfLife and qtd may be None

        # load workbooks and sheets
        sl_workbook = load_workbook(filename="data/ShelfLife.xlsx")
        q_workbook = load_workbook(filename="data/UnidadesPorCaixa.xlsx")

        sl_sheet = sl_workbook['ShelfLife']
        q_sheet = q_workbook['UnidadesPorCaixa']

        # map item id to its rows (also save the last row)
        sl_map = {}
        q_map = {}

        row = 2
        while(sl_sheet["A"+str(row)].value != None):
            sl_map[sl_sheet["A"+str(row)].value] = row
            row += 1
        slLastRow = row-1

        row = 2
        while(q_sheet["A"+str(row)].value != None):
            q_map[q_sheet["A"+str(row)].value] = row
            row += 1
        qLastRow = row-1

        # for each elm in lst:
        # try to update in ShelfLife and Quantity docs, if it exists
        # else, check if new shelflife or quantity are diff than None and add
        for elm in lst:
            if elm[0] in sl_map:
                if(elm[2] == "None" or elm[2] == ""):
                    sl_sheet["C"+str(sl_map[elm[0]])] = ""
                else:
                    sl_sheet["C"+str(sl_map[elm[0]])] = elm[2]
            if elm[0] in q_map:
                if(elm[3] == "None" or elm[3] == ""):
                    q_sheet["C"+str(q_map[elm[0]])] = ""
                else:
                    q_sheet["C"+str(q_map[elm[0]])] = elm[3]
            if elm[0] not in sl_map:
                try:
                    id = int(elm[0])
                    if(elm[2] != "None"):
                        slLastRow += 1
                        sl_sheet["A"+str(slLastRow)] = id
                        sl_sheet["B"+str(slLastRow)] = elm[1]
                        sl_sheet["C"+str(slLastRow)] = float(elm[2])
                except:
                    pass
            if elm[0] not in q_map:
                try:
                    id = int(elm[0])
                    if(elm[3] != "None"):
                        qLastRow += 1
                        q_sheet["A"+str(qLastRow)] = id
                        q_sheet["B"+str(qLastRow)] = elm[1]
                        q_sheet["C"+str(qLastRow)] = float(elm[3])
                except:
                    pass
        
        sl_workbook.save(filename="data/ShelfLife.xlsx")
        q_workbook.save(filename="data/UnidadesPorCaixa.xlsx")
